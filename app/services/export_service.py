import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display

_FONT_PATH = os.path.join(os.path.dirname(__file__), "..", "static", "fonts", "Arabic.ttf")
_FONT_NAME = "ArabicFont"
if _FONT_NAME not in pdfmetrics.getRegisteredFontNames():
    pdfmetrics.registerFont(TTFont(_FONT_NAME, _FONT_PATH))


def ar(text) -> str:
    """إعادة تشكيل النص العربي وترتيبه (RTL) ليظهر بشكل صحيح داخل PDF"""
    if text is None:
        return ""
    text = str(text)
    reshaped = arabic_reshaper.reshape(text)
    return get_display(reshaped)


def export_to_excel(headers: list, rows: list, sheet_title: str = "تقرير") -> io.BytesIO:
    """توليد ملف Excel من قائمة عناوين وصفوف بيانات"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:30]
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="0A6ED1", end_color="0A6ED1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 12), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_to_pdf(title: str, headers: list, rows: list, company_name: str = "") -> io.BytesIO:
    """توليد ملف PDF لجدول تقرير"""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm,
                             topMargin=1 * cm, bottomMargin=1 * cm)
    title_style = ParagraphStyle("ArTitle", fontName=_FONT_NAME, fontSize=16, alignment=1)
    heading_style = ParagraphStyle("ArHeading", fontName=_FONT_NAME, fontSize=12, alignment=1)
    elements = []

    if company_name:
        elements.append(Paragraph(ar(company_name), title_style))
    elements.append(Paragraph(ar(title), heading_style))
    elements.append(Spacer(1, 12))

    ar_headers = [ar(h) for h in headers]
    ar_rows = [[ar(c) if isinstance(c, str) else c for c in row] for row in rows]
    # عكس ترتيب الأعمدة ليتوافق مع اتجاه القراءة من اليمين لليسار
    ar_headers = ar_headers[::-1]
    ar_rows = [row[::-1] for row in ar_rows]

    table_data = [ar_headers] + ar_rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A6ED1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), _FONT_NAME),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6F7")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf
